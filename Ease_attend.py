import tkinter as tk
from tkinter import ttk
from openpyxl import load_workbook

def call(fl,sheet):
    win = tk.Tk()
    win.title("Easy attend")
    win.geometry("200x300")
    win.configure(bg="yellow")
    win.resizable(0,0)

    fn0=tk.StringVar()
    fn1=tk.StringVar()
    fn2=tk.StringVar()
    fn3=tk.StringVar()

    def exitt():
        wb.close()
        win.destroy()
        face()

    wb=load_workbook(fl)
    sht=wb[sheet]
    m=sht.max_row
    n=sht.max_column
    a=sht.cell(row=2,column=1).value
    b=sht.cell(row=m,column=1).value
    fn0.set(str(a)+'-'+str(b))
    fact=a-2
    from datetime import date
    today=date.today()
    def prsnt():
        try:
            m=sht.max_row
            n=sht.max_column
            sht.cell(row=1,column=n+1).value=today.strftime("%d/%m/%y")
            l1=list(map(int,fn0.get().split(',')))
            for i in range(2,m+1):
                if (i+fact) not in l1:
                    sht.cell(row=i,column=(n+1)).value = sht.cell(row=i,column=n).value + 1
                else:
                    sht.cell(row=i,column=(n+1)).value = sht.cell(row=i,column=n).value + 0
            wb.save(fl)
        except:
            from tkinter.messagebox import showinfo
            showinfo('Error','Insufficient data!')

    def absnt():
        try:
            m=sht.max_row
            n=sht.max_column
            sht.cell(row=1,column=n+1).value=today.strftime("%d/%m/%y")
            l1=list(map(int,fn0.get().split(',')))
            for i in range(2,m+1):
                if (i+fact) not in l1:
                    sht.cell(row=i,column=(n+1)).value = sht.cell(row=i,column=n).value + 0
                else:
                    sht.cell(row=i,column=(n+1)).value = sht.cell(row=i,column=n).value + 1
            wb.save(fl)
        except:
            from tkinter.messagebox import showinfo
            showinfo('Error','Insufficient data!')

    def search():
        flag=0
        m=sht.max_row
        n=sht.max_column
        for i in range(4,n+1):
            if sht.cell(row=1,column=i).value == fn1.get():
                sc = int(fn2.get())-fact
                if sht.cell(row=sc,column=i).value == sht.cell(row=sc,column=i-1).value:
                    fn3.set("Absent")
                else:
                    fn3.set('Present')
                flag=1
        if flag==0:
            from tkinter.messagebox import showinfo
            showinfo('Error','Found Nothing!')

    def reprsnt():
        try:
            m=sht.max_row
            n=sht.max_column
            sc = int(fn2.get())-fact
            if sht.cell(row=sc,column=n).value == sht.cell(row=sc,column=n-1).value:
                sht.cell(row=sc,column=n).value = sht.cell(row=sc,column=n).value + 1
            wb.save(fl)
        except:
            from tkinter.messagebox import showinfo
            showinfo('Error','Missing roll numbers!')

    def reabsnt():
        try:
            m=sht.max_row
            n=sht.max_column
            sc = int(fn2.get())-fact
            if sht.cell(row=sc,column=n).value-1 == sht.cell(row=sc,column=n-1).value:
                sht.cell(row=sc,column=n).value = sht.cell(row=sc,column=n).value - 1
            wb.save(fl)
        except:
            from tkinter.messagebox import showinfo
            showinfo('Error','Missing roll numbers!')

    stl=ttk.Style()
    stl.theme_use('clam')
    stl.configure('P.TButton',font=('Times New Roman',8,'bold'),background='green',foreground='white',width=10,borderwidth=0.5,focusthickness=3,focuscolor='none')
    stl.configure('Q.TButton',font=('Times New Roman',8,'bold'),background='brown',foreground='white',width=10,borderwidth=0.5,focusthickness=3,focuscolor='none')
    stl.configure('R.TButton',font=('Times New Roman',8,'bold'),background='orange',foreground='white',width=9,borderwidth=0.5,focusthickness=3,focuscolor='none')
    stl.configure('S.TButton',font=('Times New Roman',8,'bold'),background='green',foreground='white',width=10,borderwidth=0.5,focusthickness=3,focuscolor='none')
    stl.configure('T.TButton',font=('Times New Roman',8,'bold'),background='brown',foreground='white',width=10,borderwidth=0.5,focusthickness=3,focuscolor='none')
    stl.configure('U.TButton',font=('Times New Roman',8,'bold'),background='white',foreground='black',width=10,borderwidth=0.5,focusthickness=3,focuscolor='none')
    tk.Label(win,text='Academy Of Technology',fg='dark blue',bg='yellow',relief=tk.SOLID,font=('Times New Roman',13,'bold')).place(x=7,y=5)
    tk.Label(win,text='Roll numbers except',fg='dark blue',bg='yellow',font=('Times New Roman',8,'italic')).place(x=45,y=30)

    ttk.Button(win,text='all Present',command=prsnt,style='P.TButton').place(x=10,y=80)
    ttk.Button(win,text='all Absent',command=absnt,style='Q.TButton').place(x=106,y=80)
    ttk.Button(win,text='Search',command=search,style='R.TButton').place(x=10,y=156)
    ttk.Button(win,text='RePresent',command=reprsnt,style='S.TButton').place(x=10,y=190)
    ttk.Button(win,text='ReAbsent',command=reabsnt,style='T.TButton').place(x=105,y=190)
    
    tk.Label(win,text='Date(DD/MM/YY)',fg='dark blue',bg='yellow',font=('Times New Roman',8,'italic')).place(x=10,y=110)
    tk.Label(win,text='Roll number',fg='dark blue',bg='yellow',font=('Times New Roman',8,'italic')).place(x=115,y=110)

    ttk.Entry(win,textvariable=fn1,width=15,justify=tk.CENTER).place(x=10,y=130)
    ttk.Entry(win,textvariable=fn2,width=10,justify=tk.CENTER).place(x=115,y=130)
    
    ttk.Entry(win,textvariable=fn0,width=28,justify=tk.CENTER).place(x=10,y=50)
    ttk.Entry(win,textvariable=fn3,width=15,state='disabled',justify=tk.CENTER).place(x=85,y=160)
    ttk.Button(win,text='Exit',command=exitt,style='U.TButton').place(x=68,y=224)

    tk.Label(win,text='--by Ashirwad TechnoCraft\n\tAlik Dey',fg='dark blue',bg='yellow',font=('Times New Roman',8,'italic')).place(x=40,y=254)
        
    win.mainloop()

    

def face():
    main = tk.Tk()
    main.title("Easy attend")
    main.geometry("200x200")
    main.configure(bg="white")
    main.resizable(0,0)

    x=tk.StringVar()
    l=[]
    fl=''

    def mid():
        global fl
        main.destroy()
        call(fl,x.get())

    def brows():
        global fl
        from tkinter.filedialog import askopenfilename
        fl = askopenfilename(filetypes=[('Excel files','*.xlsx')])
        try:
            wb = load_workbook(fl)
            l=wb.sheetnames
            c1.config(value=l)
            wb.close()
        except:
            from tkinter.messagebox import showinfo
            showinfo('Error','File not choosen!')

    def cancel():
        from tkinter.messagebox import askquestion
        res=askquestion('Exit Application','Do you want to exit?')
        if(res=='yes'):
            main.destroy()

    tk.Label(main,text="Choose Folder:",bg='white',fg='dark blue',font=('Times New Roman',10,'bold')).place(x=25,y=10)

    st=ttk.Style()
    st.theme_use('clam')
    st.configure('w.TButton',font=('Times New Roman',6,'bold'),background='grey',foreground='white',width=10,borderwidth=0.5,focusthickness=3,focuscolor='none')
    ttk.Button(main,text='Browse',command=brows,style='w.TButton').place(x=120,y=10,height=25)
    
    st.configure('M.TButton',font=('Times New Roman',8,'bold'),background='green',foreground='white',width=10,borderwidth=1,focusthickness=3,focuscolor='none')
    st.configure('N.TButton',font=('Times New Roman',8,'bold'),background='brown',foreground='white',width=10,borderwidth=1,focusthickness=3,focuscolor='none')
    ttk.Button(main,text='Enter',command=mid,style='M.TButton').place(x=25,y=70)
    ttk.Button(main,text='Cancel',command=cancel,style='N.TButton').place(x=103,y=70)
    
    c1 = ttk.Combobox(main,width=22,height=4,textvariable=x,value=l)
    c1.set('Select Sheet')
    c1.place(x=25,y=40)

    lb=tk.Label(main,text='',fg='red',bg='white',font=('Times New Roman',6,'italic')).place(x=40,y=140)
    tk.Label(main,text='--by Ashirwad TechnoCraft\n\tAlik Dey',fg='black',bg='white',font=('Times New Roman',8,'italic')).place(x=40,y=160)
    
    main.mainloop()

face()
